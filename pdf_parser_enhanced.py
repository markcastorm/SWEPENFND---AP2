"""
AP2 PDF Parser - Enhanced with Advanced Table Extraction
Uses PyMuPDF (fitz) and Camelot for accurate table extraction
Extracts balance sheet data with proper column alignment
"""

import os
import glob
import pandas as pd
import fitz  # PyMuPDF
import camelot
from datetime import datetime
import re
import logging

import config

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

print("=" * 80)
print("AP2 PDF Parser - Enhanced Table Extraction Version")
print("=" * 80)


def find_latest_download_folder():
    """Find the most recent download folder"""
    download_folders = glob.glob('downloads/*')
    if not download_folders:
        logger.error("No download folders found")
        return None

    latest_folder = max(download_folders, key=os.path.getmtime)
    logger.info(f"Processing folder: {latest_folder}")
    return latest_folder


def extract_year_from_filename(filename):
    """Extract year from various filename patterns"""
    basename = os.path.basename(filename)

    patterns = [
        r'(\d{4})',  # Any 4 digits
        r'AP2_(\d{4})',  # AP2_YYYY
        r'Half.*?(\d{4})',  # Half-year-Report-YYYY
    ]

    for pattern in patterns:
        match = re.search(pattern, basename)
        if match:
            year = int(match.group(1))
            if 2000 <= year <= 2030:
                return year

    return datetime.now().year


def find_balance_sheet_page_fitz(pdf_path):
    """Find balance sheet page using PyMuPDF"""
    doc = fitz.open(pdf_path)
    best_page = None
    best_score = 0

    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text().lower()

        score = 0

        # Score based on keywords
        if 'balance sheet' in text:
            score += 30
        if 'sek million' in text:
            score += 15
        if 'total assets' in text and 'liabilities' in text:
            score += 20
        if 'fund capital' in text:
            score += 15
        if 'listed' in text and 'unlisted' in text:
            score += 10

        # Penalize summary pages
        if 'key ratios' in text:
            score -= 15
        if 'ten-year performance' in text:
            score -= 15
        if 'income statement' in text and 'balance sheet' not in text:
            score -= 10

        if score > best_score:
            best_score = score
            best_page = page_num + 1  # 1-indexed

    doc.close()

    if best_page and best_score >= 30:
        return best_page

    return None


def clean_number_string(value_str, allow_decimal=False):
    """Clean and convert Swedish formatted numbers
    Handles: '184 676', '-2 410', '464970', '458.0', '-2.4', etc.

    Args:
        value_str: String containing the number
        allow_decimal: If True, returns float for decimal numbers, else int
    """
    if pd.isna(value_str) or value_str is None:
        return None

    # Convert to string and clean
    value_str = str(value_str).strip()

    if not value_str or value_str == '' or value_str == '-':
        return None

    # Remove spaces and non-breaking spaces, keep decimal points
    cleaned = value_str.replace(' ', '').replace(',', '').replace('\xa0', '')

    # Check if it has a decimal point
    has_decimal = '.' in cleaned

    try:
        if has_decimal and allow_decimal:
            return float(cleaned)
        elif has_decimal:
            # Convert to float then to int (for values like "458.0")
            return int(float(cleaned))
        else:
            return int(cleaned)
    except ValueError:
        return None


def extract_balance_sheet_with_camelot(pdf_path, page_num):
    """Extract balance sheet table using Camelot"""
    logger.info(f"  Attempting Camelot extraction on page {page_num}...")

    try:
        # Try lattice mode first (for tables with borders)
        tables = camelot.read_pdf(
            pdf_path,
            pages=str(page_num),
            flavor='lattice',
            line_scale=40,
            strip_text='\n'
        )

        if len(tables) == 0:
            # Fallback to stream mode (for tables without borders)
            logger.info("  Lattice mode found no tables, trying stream mode...")
            tables = camelot.read_pdf(
                pdf_path,
                pages=str(page_num),
                flavor='stream',
                edge_tol=50,
                row_tol=10,
                strip_text='\n'
            )

        if len(tables) == 0:
            logger.warning("  Camelot found no tables")
            return None

        logger.info(f"  Camelot found {len(tables)} table(s)")

        # Find the best table (largest one with balance sheet structure)
        best_table = None
        best_score = 0

        for i, table in enumerate(tables):
            df = table.df
            score = 0

            # Score based on table characteristics
            if df.shape[0] > 10:  # Balance sheet has many rows
                score += 10
            if df.shape[1] >= 3:  # At least 3 columns (label + dates)
                score += 10

            # Check if it contains balance sheet keywords
            text_content = ' '.join(df.astype(str).values.flatten()).lower()
            if 'listed' in text_content:
                score += 15
            if 'assets' in text_content:
                score += 10
            if 'liabilities' in text_content:
                score += 10
            if 'fund capital' in text_content:
                score += 10

            if score > best_score:
                best_score = score
                best_table = df

        if best_table is not None and best_score >= 30:
            logger.info(f"  Selected table with score {best_score}, shape: {best_table.shape}")
            return best_table

        logger.warning(f"  Best table score {best_score} too low (threshold 30)")
        return None

    except Exception as e:
        logger.error(f"  Camelot extraction failed: {e}")
        return None


def extract_balance_sheet_with_pymupdf(pdf_path, page_num):
    """Extract balance sheet using PyMuPDF table detection"""
    logger.info(f"  Attempting PyMuPDF extraction on page {page_num}...")

    try:
        doc = fitz.open(pdf_path)
        page = doc[page_num - 1]  # 0-indexed

        # Extract tables using PyMuPDF
        tables = page.find_tables()

        if not tables or len(tables.tables) == 0:
            logger.warning("  PyMuPDF found no tables")
            doc.close()
            return None

        logger.info(f"  PyMuPDF found {len(tables.tables)} table(s)")

        # Find balance sheet table
        best_table = None
        best_score = 0

        for i, table in enumerate(tables.tables):
            df = table.to_pandas()
            score = 0

            # Score based on characteristics
            if df.shape[0] > 10:
                score += 10
            if df.shape[1] >= 3:
                score += 10

            # Check keywords
            text_content = ' '.join(df.astype(str).values.flatten()).lower()
            if 'listed' in text_content:
                score += 15
            if 'assets' in text_content:
                score += 10
            if 'liabilities' in text_content:
                score += 10
            if 'fund capital' in text_content:
                score += 10

            if score > best_score:
                best_score = score
                best_table = df

        doc.close()

        if best_table is not None and best_score >= 30:
            logger.info(f"  Selected table with score {best_score}, shape: {best_table.shape}")
            return best_table

        logger.warning(f"  Best table score {best_score} too low (threshold 30)")
        return None

    except Exception as e:
        logger.error(f"  PyMuPDF extraction failed: {e}")
        return None


def extract_key_ratios(pdf_path):
    """Extract Key Ratios data (Fund capital carried forward, Net outflows, Net result)

    Returns dict with:
        - FUNDCAPITALCARRIEDFORWARDLEVEL: Fund capital in billions (458.0)
        - NETOUTFLOWSTOTHENATIONALPENSIONSYSTEM: Net outflows in billions (-2.4)
        - TOTAL: Net result for the year in billions (1.6)
    """
    logger.info("\n  Extracting Key Ratios data...")

    try:
        # Find Key Ratios page
        doc = fitz.open(pdf_path)
        key_ratios_page = None

        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text().lower()

            if 'key ratios' in text or 'key ratio' in text:
                key_ratios_page = page_num + 1
                logger.info(f"  Found Key Ratios on page {key_ratios_page}")
                break

        doc.close()

        if not key_ratios_page:
            logger.warning("  Key Ratios page not found")
            return {}

        # Extract table using Camelot
        tables = camelot.read_pdf(pdf_path, pages=str(key_ratios_page), flavor='lattice')

        if len(tables) == 0:
            logger.info("  Lattice mode found no tables, trying stream mode...")
            tables = camelot.read_pdf(
                pdf_path,
                pages=str(key_ratios_page),
                flavor='stream',
                edge_tol=50,
                row_tol=10
            )

        if len(tables) == 0:
            logger.warning("  No tables found on Key Ratios page")
            return {}

        df = tables[0].df
        logger.info(f"  Key Ratios table shape: {df.shape}")

        # Extract the 3 values we need
        data = {}

        for idx, row in df.iterrows():
            field_name = str(row.iloc[0]).strip().lower() if not pd.isna(row.iloc[0]) else ""

            # Fund capital carried forward (LEVEL - in billions)
            if 'fund capital carried forward' in field_name and 'sek billion' in field_name:
                value = clean_number_string(row.iloc[1], allow_decimal=True)
                if value is not None:
                    data['FUNDCAPITALCARRIEDFORWARDLEVEL'] = value
                    logger.info(f"    [OK] Fund capital carried forward (LEVEL): {value}")

            # Net result for the period (in billions)
            elif 'net result for the period' in field_name and 'sek billion' in field_name:
                value = clean_number_string(row.iloc[1], allow_decimal=True)
                if value is not None:
                    data['TOTAL'] = value
                    logger.info(f"    [OK] Net result for the year (TOTAL): {value}")

            # Net outflows to national pension system (in billions)
            elif 'net outfl' in field_name and 'national pension' in field_name:
                value = clean_number_string(row.iloc[1], allow_decimal=True)
                if value is not None:
                    data['NETOUTFLOWSTOTHENATIONALPENSIONSYSTEM'] = value
                    logger.info(f"    [OK] Net outflows to pension system: {value}")

        logger.info(f"    [INFO] Extracted {len(data)}/3 Key Ratios fields")
        return data

    except Exception as e:
        logger.error(f"  Key Ratios extraction failed: {e}")
        return {}


def parse_balance_sheet_from_table(df):
    """Parse balance sheet data from extracted table DataFrame

    Expected structure:
    Column 0: Field names (Listed, Unlisted, etc.)
    Column 1: Latest period (30 Jun 2025)
    Column 2: Prior period (30 Jun 2024)
    Column 3: Year end (31 Dec 2024)

    We want column 1 (latest period)
    """
    if df is None or df.empty:
        return {}

    logger.info(f"  Parsing table with shape: {df.shape}")

    # Field mapping patterns - flexible to match table variations
    # IMPORTANT: Check UNLISTED before LISTED to avoid substring matches
    field_patterns = {
        'EQUITIESANDPARTICIPATIONSUNLISTED': [r'^\s*unlisted\s*$', r'non-listed', r'equities.*unlisted'],
        'EQUITIESANDPARTICIPATIONSLISTED': [r'^\s*listed\s*$', r'(?<!un)listed', r'equities.*listed'],
        'BONDSANDOTHERFIXEDINCOMESECURITIES': [r'bonds.*fixed.?income', r'fixed.?income.*securities'],
        'DERIVATIVEINSTRUMENTS': [r'^\s*derivative instruments\s*$'],
        'CASHANDBANKBALANCES': [r'cash.*bank', r'cash and bank balances'],
        'OTHERASSETS': [r'^\s*other assets\s*$', r'other assets(?!\w)'],
        'PREPAIDEXPENSESANDACCRUEDINCOME': [r'prepaid.*accrued.*income'],
        'TOTALASSETS': [r'^\s*total\s+assets\s*$', r'total assets(?!\w)'],
        'DERIVATIVEINSTRUMENTSLIABILITIES': [r'^\s*derivative instruments\s*$'],  # Context-aware
        'OTHERLIABILITIES': [r'^\s*other liabilities\s*$', r'other liabilities(?!\w)'],
        'DEFERREDINCOMEANDACCRUEDEXPENSES': [r'deferred.*accrued.*expenses', r'deferred income.*accrued'],
        'TOTALLIABILITIES': [r'^\s*total liabilities\s*$', r'total liabilities(?!\w)'],
        'FUNDCAPITALCARRIEDFORWARD': [r'fund capital carried forward', r'carried.*forward'],
        'NETPAYMENTSTOTHENATIONALPENSIONSYSTEM': [r'net.*national pension', r'net payments.*pension'],
        'NETRESULTFORTHEPERIOD': [r'net result.*period'],
        'TOTALFUNDCAPITAL': [r'^\s*total fund capital\s*$', r'total fund capital(?!\w)'],
        'TOTALFUNDCAPITALANDLIABILITIES': [r'total fund capital and liabilities', r'total.*capital.*liabilities']
    }

    data = {}
    in_assets_section = True  # Start in assets
    assets_found = False
    liabilities_found = False

    # Iterate through rows
    for idx, row in df.iterrows():
        # Get first column as field name
        field_name = str(row.iloc[0]).strip().lower() if not pd.isna(row.iloc[0]) else ""

        if not field_name:
            continue

        # Track sections - only for section headers, not data rows
        if field_name == 'assets':  # Section header
            in_assets_section = True
            assets_found = True
            continue
        elif field_name == 'liabilities' or field_name == 'fund capital and liabilities':  # Section headers
            in_assets_section = False
            liabilities_found = True
            continue
        elif field_name == 'fund capital':  # Section header
            in_assets_section = False
            continue

        # Try to match field patterns
        field_matched = False
        for field_key, patterns in field_patterns.items():
            # Skip if already extracted this field
            if field_key in data:
                continue

            matched = False
            for pattern in patterns:
                if re.search(pattern, field_name, re.IGNORECASE):
                    matched = True
                    break

            if not matched:
                continue

            # Context validation for derivative instruments
            if field_key == 'DERIVATIVEINSTRUMENTS' and not in_assets_section:
                continue
            if field_key == 'DERIVATIVEINSTRUMENTSLIABILITIES' and in_assets_section:
                continue

            # Extract value from column 1 (latest period)
            if df.shape[1] >= 2:
                value = clean_number_string(row.iloc[1])

                if value is not None:
                    data[field_key] = value
                    logger.info(f"    [OK] {field_key}: {value:,}")
                    field_matched = True
                    break  # Break out of field_patterns loop

        # Move to next row if we matched this row
        if field_matched:
            continue

    # Summary
    logger.info(f"    [INFO] Extracted {len(data)}/17 fields")

    return data


def parse_balance_sheet_adaptive(pdf_path, year):
    """Parse balance sheet using multiple extraction methods"""
    logger.info(f"\nProcessing: {os.path.basename(pdf_path)} (Year: {year})")

    # Step 1: Find balance sheet page
    page_num = find_balance_sheet_page_fitz(pdf_path)

    if not page_num:
        logger.error("  Could not find balance sheet page")
        return {}

    logger.info(f"  Found balance sheet on page {page_num}")

    # Step 2: Try Camelot extraction first
    df = extract_balance_sheet_with_camelot(pdf_path, page_num)

    # Step 3: Fallback to PyMuPDF if Camelot fails
    if df is None:
        logger.info("  Trying PyMuPDF as fallback...")
        df = extract_balance_sheet_with_pymupdf(pdf_path, page_num)

    if df is None:
        logger.error("  All table extraction methods failed")
        return {}

    # Step 4: Parse data from table
    data = parse_balance_sheet_from_table(df)

    # Step 5: Extract Key Ratios data
    key_ratios_data = extract_key_ratios(pdf_path)

    # Step 6: Merge Key Ratios into main data
    data.update(key_ratios_data)

    # Step 7: Validate
    if data:
        validate_balance_sheet(data)

    logger.info(f"\n  [FINAL] Extracted {len(data)}/20 total fields (17 balance sheet + 3 key ratios)")

    return data


def validate_balance_sheet(data):
    """Validate extracted balance sheet data"""
    validations_passed = 0
    total_validations = 0

    # Check 1: Assets = Fund Capital + Liabilities
    if 'TOTALASSETS' in data and 'TOTALFUNDCAPITALANDLIABILITIES' in data:
        total_validations += 1
        diff = abs(data['TOTALASSETS'] - data['TOTALFUNDCAPITALANDLIABILITIES'])
        if diff <= 100:
            logger.info(f"    [OK] Balance sheet balances (diff: {diff})")
            validations_passed += 1
        else:
            logger.warning(f"    [WARN] Assets ({data['TOTALASSETS']:,}) != Fund+Liab ({data['TOTALFUNDCAPITALANDLIABILITIES']:,})")

    # Check 2: Fund Capital components
    if all(k in data for k in ['TOTALFUNDCAPITAL', 'TOTALLIABILITIES', 'TOTALFUNDCAPITALANDLIABILITIES']):
        total_validations += 1
        expected = data['TOTALFUNDCAPITAL'] + data['TOTALLIABILITIES']
        actual = data['TOTALFUNDCAPITALANDLIABILITIES']
        if abs(expected - actual) <= 100:
            logger.info(f"    [OK] Fund + Liabilities = Total")
            validations_passed += 1
        else:
            logger.warning(f"    [WARN] Fund({data['TOTALFUNDCAPITAL']:,}) + Liab({data['TOTALLIABILITIES']:,}) != Total({actual:,})")

    if total_validations > 0:
        logger.info(f"    [INFO] Validation: {validations_passed}/{total_validations} checks passed")


def create_output(all_data):
    """Create Excel output matching sample structure"""
    logger.info(f"\n{'='*80}")
    logger.info("CREATING OUTPUT")
    logger.info(f"{'='*80}")

    # Create DataFrame
    df_data = []

    for year, data in all_data.items():
        row = {'Unnamed: 0': year}

        # Map extracted fields to output headers
        for header in config.OUTPUT_HEADERS[1:]:
            value = None

            # Extract field name from header
            # e.g., 'AP2.EQUITIESANDPARTICIPATIONSLISTED.FLOW.NONE.H.1@AP2' -> 'EQUITIESANDPARTICIPATIONSLISTED'
            # e.g., 'AP2.FUNDCAPITALCARRIEDFORWARD.LEVEL.NONE.H.1@AP2' -> 'FUNDCAPITALCARRIEDFORWARDLEVEL'
            parts = header.split('.')
            if len(parts) >= 2:
                # Check if this is the special LEVEL column
                if len(parts) >= 3 and parts[2] == 'LEVEL':
                    field_name = parts[1] + 'LEVEL'
                else:
                    field_name = parts[1]

                value = data.get(field_name)

            row[header] = value

        df_data.append(row)

    df = pd.DataFrame(df_data, columns=config.OUTPUT_HEADERS)

    # Create output folders
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_folder = os.path.join('output', timestamp)
    latest_folder = os.path.join('output', 'latest')

    os.makedirs(output_folder, exist_ok=True)
    os.makedirs(latest_folder, exist_ok=True)

    # Save files with 2 header rows (technical + sub-headers)
    output_file = os.path.join(output_folder, f'AP2_Financial_Data_{timestamp}.xlsx')
    latest_file = os.path.join(latest_folder, 'AP2_Financial_Data_latest.xlsx')

    # Save with custom header structure using openpyxl
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows

    for file_path in [output_file, latest_file]:
        wb = openpyxl.Workbook()
        ws = wb.active

        # Row 1: Technical headers (first column should be None, not 'Unnamed: 0')
        for col_idx, header in enumerate(config.OUTPUT_HEADERS, start=1):
            value = None if header == 'Unnamed: 0' else header
            ws.cell(row=1, column=col_idx, value=value)

        # Row 2: Human-readable sub-headers
        for col_idx, subheader in enumerate(config.OUTPUT_SUBHEADERS, start=1):
            ws.cell(row=2, column=col_idx, value=subheader)

        # Row 3+: Data
        for row_idx, (_, row_data) in enumerate(df.iterrows(), start=3):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(file_path)

    logger.info(f"✓ Saved: {output_file}")
    logger.info(f"✓ Saved: {latest_file}")

    # Calculate accuracy
    filled_count = df.notna().sum().sum() - len(df)
    total_cells = len(df) * (len(config.OUTPUT_HEADERS) - 1)
    accuracy = (filled_count/total_cells*100) if total_cells > 0 else 0

    logger.info(f"✓ Filled {filled_count}/{total_cells} cells ({accuracy:.1f}%)")

    if accuracy >= 90:
        logger.info("[EXCELLENT] High extraction accuracy!")
    elif accuracy >= 70:
        logger.info("[GOOD] Acceptable extraction accuracy")
    else:
        logger.warning("[POOR] Low extraction accuracy")

    return output_file


def main():
    """Main execution"""
    try:
        # Find latest download folder
        download_folder = find_latest_download_folder()
        if not download_folder:
            return

        # Find PDFs
        pdf_files = glob.glob(os.path.join(download_folder, '*.pdf'))
        logger.info(f"Found {len(pdf_files)} PDF file(s)\n")

        if not pdf_files:
            logger.error("No PDF files found")
            return

        # Process each PDF
        all_data = {}

        for pdf_file in pdf_files:
            year = extract_year_from_filename(pdf_file)
            data = parse_balance_sheet_adaptive(pdf_file, year)

            if data:
                all_data[year] = data
                logger.info(f"  ✓ Extracted {len(data)} fields for {year}")
            else:
                logger.warning(f"  ⚠ No data extracted for {year}")

        if not all_data:
            logger.error("\n[ERROR] No financial data extracted")
            return

        # Create output
        output_file = create_output(all_data)

        logger.info(f"\n{'='*80}")
        logger.info("ENHANCED PARSER COMPLETED")
        logger.info(f"{'='*80}")
        logger.info(f"✓ Processed {len(pdf_files)} PDF(s)")
        logger.info(f"✓ Extracted data for {len(all_data)} year(s): {list(all_data.keys())}")
        logger.info(f"✓ Output: {output_file}")
        logger.info(f"{'='*80}")

    except Exception as e:
        logger.error(f"\n[FATAL ERROR] {e}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    main()
