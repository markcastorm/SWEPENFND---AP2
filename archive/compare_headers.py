import openpyxl

our = openpyxl.load_workbook('output/latest/AP2_Financial_Data_latest.xlsx')
sample = openpyxl.load_workbook('project information/AP2_SA_SWEPENFND_DATA_20220920.xlsx')

print('Our Row 1, Cell 1:', repr(our.active.cell(1,1).value))
print('Sample Row 1, Cell 1:', repr(sample.active.cell(1,1).value))

print('\nFull Row 1 comparison:')
for i in range(1, 22):
    our_val = our.active.cell(1,i).value
    sample_val = sample.active.cell(1,i).value
    match = our_val == sample_val
    if not match:
        print(f'  Col {i}: MISMATCH')
        print(f'    Our: {repr(our_val)}')
        print(f'    Sample: {repr(sample_val)}')

print('\nRow 2 comparison (first 5):')
for i in range(1, 6):
    our_val = our.active.cell(2,i).value
    sample_val = sample.active.cell(2,i).value
    print(f'  Col {i}: Match={our_val == sample_val}')
